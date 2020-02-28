import requests, json, re, os
import hashlib
import lxml.html as html
#from lxml.html import parse, fromstring
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

cols = ['ID scan','ID','Фамилия','Имя','Отчество','Дата рождения/Возраст','Место рождения','Дата и место призыва','Последнее место службы','Воинское звание','Судьба','Дата смерти','Первичное место захоронения']

def parse_file (name_file):
    dict_ = {}
    f = open(name_file, 'r')
    s = f.read()
    dict_={}
    list_ = s.splitlines()
    for item in list_:
        items = item.split(":")
        dict_[items[0]] = items[1].lstrip()
    return dict_
#####################################
def make_str_cookie(cookies):
    str_cook = ''
    for key, value in cookies.items():
        str_cook += '{0}={1};'.format(key,value)
    return str_cook
#####################################
def getStringHash(id):
    h = hashlib.md5(str(id).encode()+b'db76xdlrtxcxcghn7yusxjcdxsbtq1hnicnaspohh5tzbtgqjixzc5nmhybeh')
    p = h.hexdigest()
    return str(p)
#####################################
def get_info(id_scan,id):
    info_url = 'https://obd-memorial.ru/html/info.htm?id='+str(id)
    res3 = requests.get(info_url,allow_redirects = True)
    doc = html.fromstring(res3.text)
    divs = {}
    for div in doc.find_class('card_parameter'):
        divs[div.getchildren()[0].text_content()] = div.getchildren()[1].text_content()
        #print ('%s: %s' % (div.getchildren()[0].text_content(), div.getchildren()[1].text_content()))
    list_col = []
    for col in cols:
        if(col in divs.keys()):
            list_col.append(divs[col])
        else:
            list_col.append('')
    list_col[1] = id
    list_col[0] = id_scan
    return list_col
#####################################

info_url = 'https://obd-memorial.ru/html/info.htm?id=70782617'
img_info = 'https://obd-memorial.ru/html/getimageinfo?id=70782617'

res1 = requests.get(info_url,allow_redirects = True)
if(res1.status_code==307):
    print(res1.status_code)
    print('*****************')

    cookies = {}
    cookies['3fbe47cd30daea60fc16041479413da2']=res1.cookies['3fbe47cd30daea60fc16041479413da2']
    cookies['JSESSIONID']=res1.cookies['JSESSIONID']

    response = requests.get(img_info)
    response_dict = json.loads(response.text)
    #print(response_dict[0]['mapData'].keys()[0])
    i=0
    row_num = 1
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Person'
    columns = cols
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for item in response_dict:
        i+=1
        print(i, item['id'])
        img_url="https://obd-memorial.ru/html/images3?id="+str(item['id'])+"&id1="+(getStringHash(item['id']))+"&path="+item['img']
        headers_302 = parse_file(BASE_DIR+'/header_302.txt')
        headers_302['Cookie'] = make_str_cookie(cookies)
        headers_302['Referer'] = info_url
        req302 = requests.get(img_url,headers=headers_302,cookies=cookies, allow_redirects = False)
        if(req302.status_code==302):
            params = {}
            params['id'] = str(item['id'])
            params['id1'] = getStringHash(item['id'])
            params['path'] = item['img']
            headers_img = parse_file(BASE_DIR+'/header_img.txt')
            headers_img['Referer'] = info_url
            #####################
            req_img = requests.get("https://cdn.obd-memorial.ru/html/images3",headers=headers_img,params=params,stream = True,allow_redirects = False )
            #####################
            if(req_img.status_code==200):
                for id in item['mapData'].keys():
                    row_num += 1
                    row = get_info(item['id'],id)
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value


                #location = os.path.abspath("./scan/"+str(item['id'])+'.jpg')
                #f = open(location, 'wb')
                #f.write(req_img.content)
                #f.close()
        
    workbook.save(filename = 'sample_book.xlsx')
    exit(1)

