import requests, json, re, os
import hashlib
import lxml.html as html
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload,MediaFileUpload
from googleapiclient.discovery import build
from string import Template
import pprint
import io
from pathlib import Path

#from lxml.html import parse, fromstring
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

cols = ['ID scan','ID','Фамилия','Имя','Отчество','Дата рождения/Возраст','Место рождения','Дата и место призыва','Последнее место службы','Воинское звание','Судьба','Дата смерти','Первичное место захоронения']

pp = pprint.PrettyPrinter(indent=4)
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'obd.json'
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('drive', 'v3', credentials=credentials)
name_root_folder = 'Folder'
tmpFolder='tmp'
Path(tmpFolder).mkdir(parents=True, exist_ok=True)
root_results = service.files().list(pageSize=10,fields="nextPageToken, files(id, name, mimeType)",q=Template("name contains '$name_root_folder'").safe_substitute(name_root_folder=name_root_folder)).execute()
#pp.pprint(root_results)
id_root_folder = root_results['files'][0]['id']

######################################
def parse_file (name_file):
    dict_ = {}
    f = open(name_file, 'r')
    s = f.read()
    dict_={}
    list_ = s.splitlines()
    for item in list_:
        items = item.split(":")
        dict_[items[0]] = items[1].lstrip()
    return dict_
#####################################
def make_str_cookie(cookies):
    str_cook = ''
    for key, value in cookies.items():
        str_cook += '{0}={1};'.format(key,value)
    return str_cook
#####################################
def getStringHash(id):
    h = hashlib.md5(str(id).encode()+b'db76xdlrtxcxcghn7yusxjcdxsbtq1hnicnaspohh5tzbtgqjixzc5nmhybeh')
    p = h.hexdigest()
    return str(p)
#####################################
def get_info(id_scan,id,cookies):
    cookies['showimage']='0'
    info_url = 'https://obd-memorial.ru/html/info.htm?id='+str(id)
    res3 = requests.get(info_url,cookies=cookies)
    doc = html.fromstring(res3.text)
    divs = {}
    for div in doc.find_class('card_parameter'):
        divs[div.getchildren()[0].text_content()] = div.getchildren()[1].text_content()
        #print ('%s: %s' % (div.getchildren()[0].text_content(), div.getchildren()[1].text_content()))
    list_col = []
    for col in cols:
        if(col in divs.keys()):
            list_col.append(divs[col])
        else:
            list_col.append('')
    list_col[1] = id
    list_col[0] = id_scan
    return list_col
#####################################

def main(id,image,excel):
    info_url = 'https://obd-memorial.ru/html/info.htm?id={}'.format(id)
    img_info = 'https://obd-memorial.ru/html/getimageinfo?id={}'.format(id)
    res1 = requests.get(info_url,allow_redirects = True)
    if(res1.status_code==307):
        print(res1.status_code)
        print('*****************')

        cookies = {}
        cookies['3fbe47cd30daea60fc16041479413da2']=res1.cookies['3fbe47cd30daea60fc16041479413da2']
        cookies['JSESSIONID']=res1.cookies['JSESSIONID']

        response = requests.get(img_info)
        response_dict = json.loads(response.text)
        #print(response_dict[0]['mapData'].keys()[0])
        i=0
        if(excel):
            row_num = 1
            workbook = Workbook()
            # Get active worksheet/tab
            worksheet = workbook.active
            worksheet.title = 'Person'
            columns = cols
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
        # идем по списку id сканов
        # сохраняем имена файлов [id].jpg в list, потом его вернем для выгрузки Google Drive
        list_file = []
        for item in response_dict:
            i+=1
            print(i, item['id'])

            if(excel):
                for id in item['mapData'].keys():
                    row_num += 1
                    row = get_info(item['id'],id,cookies)
                    #print('\t',id)
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value


            if(image):
                img_url="https://obd-memorial.ru/html/images3?id="+str(item['id'])+"&id1="+(getStringHash(item['id']))+"&path="+item['img']
                headers_302 = parse_file(BASE_DIR+'/header_302.txt')
                headers_302['Cookie'] = make_str_cookie(cookies)
                headers_302['Referer'] = info_url
                req302 = requests.get(img_url,headers=headers_302,cookies=cookies, allow_redirects = False)
                if(req302.status_code==302):
                    params = {}
                    params['id'] = str(item['id'])
                    params['id1'] = getStringHash(item['id'])
                    params['path'] = item['img']
                    headers_img = parse_file(BASE_DIR+'/header_img.txt')
                    headers_img['Referer'] = info_url
                    #####################
                    req_img = requests.get("https://cdn.obd-memorial.ru/html/images3",headers=headers_img,params=params,stream = True,allow_redirects = False )
                    #####################
                    if(req_img.status_code==200):
                        location = os.path.abspath("./scan/"+str(item['id'])+'.jpg')
                        f = open(location, 'wb')
                        f.write(req_img.content)
                        f.close()
                        list_file.append("./scan/"+str(item['id'])+'.jpg')
    if(excel):
        workbook.save(filename = 'sample_book.xlsx')
    return list_file

#####################################
# def main() end
#####################################


def save_to_folder(name_folder_save, list_files_upload):
    result = service.files().list(pageSize=10,fields="nextPageToken, files(id, name, mimeType,webViewLink)",q=Template("name contains '$name_folder_save'").safe_substitute(name_folder_save=name_folder_save)).execute()
    if(not result['files']):
        #create catalog
        file_metadata = {
            'name': name_folder_save,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [id_root_folder]
        }
        result = service.files().create(body=file_metadata, fields='id').execute()
    # id каталога для сохранения
    id_folder_save = result['files'][0]['id']
    # ссылка на каталог
    web_link = result['files'][0]['webViewLink']

    # получим список файлов в каталоге GoogleDrive, что бы не делать лишней работы
    list_files = service.files().list(pageSize=1000,fields="nextPageToken, files(id, name, mimeType, parents)",q=Template("'$id_parents_folder' in parents").safe_substitute(id_parents_folder=id_folder_save)).execute()
    control_list_file = []
    for file in list_files['files']:
        control_list_file.append(file['name'])
    ################################################################################
    #return
    for _file in list_files_upload:
        name = os.path.basename(_file)

        if(name not in control_list_file):
            print(name)
            file_metadata = {'name': name,'parents': [id_folder_save]}
            media = MediaFileUpload(_file, resumable=True)
            r = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
            if(r['id']):
                media = None
                os.remove(_file)
    return web_link

_id = 51480906
d = {'image':True, 'excel':False}
list_file = main(_id,**d)
for name in list_file:
    name = os.path.basename(name)
    print(name)

link = save_to_folder(str(_id),list_file)
print(link)
